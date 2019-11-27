from openpyxl import load_workbook

from django.core.management.base import BaseCommand

from lcc.models import (
    TaxonomyClassification,
    Question,
    Gap,
)


class Command(BaseCommand):

    help = """
        Import Questions and Gaps from an excel file.
        File example in commands/example/original_questions.xlsx
    """

    def process(self, code):
        """
        Updates a code based on the number of previous taxonomy classifications
        sent as parameter. For example, if the questions in a file use taxonomy
        classification codes starting from 1.1, but in our database we already
        have 8 top level classifications, the returned code will be 9.1.
        """
        pt1, pt2 = code.split('.', 1)
        return '.'.join([str(int(pt1) + self.num_prev), pt2])

    def add_arguments(self, parser):
        parser.add_argument('file', type=str)
        # num_prev is the number of previous top-level taxonomy classifications
        # present in the system.
        parser.add_argument("-n", "--num_prev", type=int)

    def parse_row(self, row):
        value = row[0].value
        if value:
            raw_code, text = value.strip().split(' ', 1)
            code = self.process(raw_code)
            return {
                'text': text,
                'level': 0,
                'parent_answer': 1,
                'classification': code,
                'gap_answer': 0,
                'gap_classifications': [code],
            }
        else:
            value = row[3].value
            raw_code, text = value.strip().split(' ', 1)
            code = self.process(raw_code)
            return {
                'text': text,
                'level': 1,
                'parent_answer': 1,
                'classification': code,
                'gap_answer': 0,
                'gap_classifications': [code],
            }

    def create_question(self, data, parents_by_level):
        if data['level'] == 0:
            parent = None
        else:
            parent = parents_by_level[data['level'] - 1]
        classification = TaxonomyClassification.objects.get(
            code=data['classification']
        )
        question = Question.objects.filter(
            classification=classification,
            parent=parent
        ).first()
        if question:
            parents_by_level[data['level']] = question
            print("Question for {} already created.".format(classification))
            return
        question = Question.objects.create(
            text=data['text'],
            parent=parent,
            parent_answer=data['parent_answer'],
            classification=classification
        )
        return question

    def create_gap(self, data, question):
        gap_classifications = []
        for code in data['gap_classifications']:
            classification = TaxonomyClassification.objects.get(code=code)
            gap_classifications.append(classification)

        gap = Gap.objects.create(on=data['gap_answer'], question=question)
        for classification in gap_classifications:
            gap.classifications.add(classification)

    def handle(self, file, *args, **options):
        self.num_prev = options.get('num_prev', 0)
        wb = load_workbook(file, read_only=True)
        parents_by_level = [None] * 10
        for sheet in wb:
            for row in sheet.iter_rows(min_row=2):
                if not row or not (row[0].value or row[3].value):
                    # No relevant data in this row that cannot be inferred
                    # from other ones, skip it
                    continue
                data = self.parse_row(row)
                try:
                    question = self.create_question(data, parents_by_level)
                    if not question:
                        continue
                    parents_by_level[data['level']] = question
                    self.create_gap(data, question)
                except Exception as e:
                    print(
                        "Failed to create question for {} with error {}".format(
                            data['classification'], str(e)
                        )
                    )
